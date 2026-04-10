#!/usr/bin/env python3
"""
Sandbox Tracker - Simuleeritud kauplemise jälgimine.
Logib kõik sandbox-tehingud JSON faili, jälgib igapäevaseid hindu
ja genereerib professionaalse Excel raamatu.

KAPITALI HALDUS:
- Stardikapital on fikseeritud (vaikimisi €500)
- Ost vähendab vaba raha (available_cash)
- Müük lisab tagasi müügitulu (mitte ainult kasumit, vaid kogu summa)
- Kui raha pole, ei saa osta
- Kapital muutub kumulatiivselt vastavalt tehingutele

Erinevalt trade_journal.py-st, siin EI tehta reaalseid tehinguid Lightyearis.
Kõik on ainult simuleeritud ja logitud Excelisse.
"""
import json
import sys
import os
from datetime import datetime, timedelta
from collections import defaultdict

SANDBOX_DIR = os.path.expanduser("~/.trading-agents/sandbox")
SANDBOX_FILE = os.path.join(SANDBOX_DIR, "sandbox_journal.json")
CONFIG_FILE = os.path.join(SANDBOX_DIR, "sandbox_config.json")
PRICES_FILE = os.path.join(SANDBOX_DIR, "price_history.json")
SNAPSHOTS_FILE = os.path.join(SANDBOX_DIR, "daily_snapshots.json")
WEEKLY_FILE = os.path.join(SANDBOX_DIR, "weekly_summaries.json")

DEFAULT_CONFIG = {
    "starting_capital_eur": 500,
    "max_trade_eur": 500,
    "target_profit_pct": 10,
    "stop_loss_pct": -7,
    "max_open_positions": 5,
    "risk_level": "medium",
    "email": "<YOUR_EMAIL>",
}


def ensure_dir():
    os.makedirs(SANDBOX_DIR, exist_ok=True)


def load_json(filepath, default=None):
    ensure_dir()
    if os.path.exists(filepath):
        with open(filepath, "r") as f:
            return json.load(f)
    return default if default is not None else {}


def save_json(filepath, data):
    ensure_dir()
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2, default=str)


def load_journal():
    return load_json(SANDBOX_FILE, {"trades": [], "stats": {}, "available_cash_eur": None})


def save_journal(journal):
    save_json(SANDBOX_FILE, journal)


def load_config():
    config = load_json(CONFIG_FILE)
    if not config:
        save_json(CONFIG_FILE, DEFAULT_CONFIG)
        return DEFAULT_CONFIG
    # Ensure starting_capital exists in old configs
    if "starting_capital_eur" not in config:
        config["starting_capital_eur"] = 500
        save_json(CONFIG_FILE, config)
    return config


def calculate_available_cash():
    """Arvuta vaba raha ALATI tehingutest (mitte salvestatud väärtusest).
    See tagab et raha ei saa kunagi negatiivseks ega valeks minna.
    Stardikapital: €500 fikseeritud. Ostud vähendavad, müügid lisavad."""
    journal = load_journal()
    config = load_config()
    starting = config["starting_capital_eur"]

    # ALATI arvuta tehingutest - ära usalda salvestatud väärtust
    cash = starting
    for t in journal.get("trades", []):
        if t["action"] == "BUY":
            cash -= t["amount_eur"]
        elif t["action"] == "SELL":
            cash += t["amount_eur"]

    # Raha EI SAA olla negatiivne
    cash = max(round(cash, 2), 0)

    # Uuenda journal'i väärtus
    journal["available_cash_eur"] = cash
    save_journal(journal)
    return cash


def get_total_capital():
    """Arvuta kogukapital = vaba raha + avatud positsioonide väärtus."""
    cash = calculate_available_cash()
    open_pos = get_open_positions()
    invested = sum(t["amount_eur"] for t in open_pos)
    return round(cash + invested, 2)


# =============================================
# TEHINGUTE HALDUS
# =============================================

def add_trade(symbol, action, shares, price, amount_eur, reason, score=0):
    """Lisa uus sandbox-tehing. Kontrollib ka raha olemasolu."""
    journal = load_journal()
    config = load_config()
    amount_eur = round(float(amount_eur), 2)

    # OSTU puhul kontrolli kas piisavalt raha
    if action.upper() == "BUY":
        cash = calculate_available_cash()
        if amount_eur > cash:
            error = {
                "error": "INSUFFICIENT_FUNDS",
                "available_cash_eur": cash,
                "requested_eur": amount_eur,
                "message": f"Pole piisavalt vaba raha! Saadaval: €{cash:.2f}, küsitud: €{amount_eur:.2f}"
            }
            print(json.dumps(error, indent=2))
            return error

    trade = {
        "id": len(journal["trades"]) + 1,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "symbol": symbol.upper(),
        "action": action.upper(),
        "shares": float(shares),
        "price": float(price),
        "amount_eur": amount_eur,
        "reason": reason,
        "analysis_score": int(score),
        "status": "OPEN" if action.upper() == "BUY" else "CLOSED",
        "profit_eur": None,
        "profit_pct": None,
        "mode": "SANDBOX",
    }

    if action.upper() == "SELL":
        for t in reversed(journal["trades"]):
            if t["symbol"] == symbol.upper() and t["action"] == "BUY" and t["status"] == "OPEN":
                buy_price = t["price"]
                profit_pct = round((float(price) / buy_price - 1) * 100, 2)
                profit_eur = round(float(amount_eur) - t["amount_eur"], 2)
                trade["profit_pct"] = profit_pct
                trade["profit_eur"] = profit_eur
                trade["buy_trade_id"] = t["id"]
                t["status"] = "CLOSED"
                t["closed_date"] = trade["date"]
                t["profit_pct"] = profit_pct
                t["profit_eur"] = profit_eur
                break

    journal["trades"].append(trade)
    journal["stats"] = calculate_stats(journal["trades"])

    # Uuenda vaba raha
    if action.upper() == "BUY":
        journal["available_cash_eur"] = round(journal.get("available_cash_eur", calculate_available_cash()) - amount_eur, 2)
    elif action.upper() == "SELL":
        journal["available_cash_eur"] = round(journal.get("available_cash_eur", calculate_available_cash()) + amount_eur, 2)

    save_journal(journal)
    return trade


def calculate_stats(trades):
    """Arvuta üldine statistika."""
    closed = [t for t in trades if t.get("profit_pct") is not None and t["action"] == "SELL"]
    open_trades = [t for t in trades if t["action"] == "BUY" and t["status"] == "OPEN"]

    if not closed:
        return {
            "total_trades": len(trades),
            "open_positions": len(open_trades),
            "open_symbols": [t["symbol"] for t in open_trades],
            "closed_trades": 0,
            "win_rate": 0,
            "total_profit_eur": 0,
            "avg_profit_pct": 0,
        }

    wins = [t for t in closed if t["profit_pct"] > 0]

    return {
        "total_trades": len(trades),
        "open_positions": len(open_trades),
        "open_symbols": [t["symbol"] for t in open_trades],
        "closed_trades": len(closed),
        "wins": len(wins),
        "losses": len(closed) - len(wins),
        "win_rate": round(len(wins) / len(closed) * 100, 1),
        "total_profit_eur": round(sum(t.get("profit_eur", 0) for t in closed if t.get("profit_eur")), 2),
        "avg_profit_pct": round(sum(t["profit_pct"] for t in closed) / len(closed), 2),
        "best_trade_pct": round(max(t["profit_pct"] for t in closed), 2),
        "worst_trade_pct": round(min(t["profit_pct"] for t in closed), 2),
        "best_trade": max(closed, key=lambda t: t["profit_pct"])["symbol"],
        "worst_trade": min(closed, key=lambda t: t["profit_pct"])["symbol"],
    }


def get_open_positions():
    journal = load_journal()
    return [t for t in journal["trades"] if t["action"] == "BUY" and t["status"] == "OPEN"]


# =============================================
# HINDADE JÄLGIMINE
# =============================================

def check_prices():
    """Kontrolli avatud positsioonide hetkehindu."""
    import yfinance as yf

    positions = get_open_positions()
    config = load_config()
    results = []

    for pos in positions:
        try:
            ticker = yf.Ticker(pos["symbol"])
            current = ticker.info.get("currentPrice") or ticker.info.get("regularMarketPrice", 0)

            if current and pos["price"] > 0:
                pct_change = (current / pos["price"] - 1) * 100
                result = {
                    "symbol": pos["symbol"],
                    "buy_price": pos["price"],
                    "current_price": round(current, 2),
                    "pct_change": round(pct_change, 2),
                    "buy_date": pos["date"],
                    "trade_id": pos["id"],
                    "amount_eur": pos["amount_eur"],
                    "current_value_eur": round(pos["amount_eur"] * (1 + pct_change / 100), 2),
                    "unrealized_pnl_eur": round(pos["amount_eur"] * pct_change / 100, 2),
                }

                if pct_change >= config["target_profit_pct"]:
                    result["signal"] = "SELL_PROFIT"
                    result["signal_reason"] = f"Sihtmärk saavutatud! +{pct_change:.1f}%"
                elif pct_change <= config["stop_loss_pct"]:
                    result["signal"] = "SELL_STOPLOSS"
                    result["signal_reason"] = f"Stop-loss! {pct_change:.1f}%"
                else:
                    result["signal"] = "HOLD"
                    result["signal_reason"] = f"Hoiame ({pct_change:+.1f}%)"

                results.append(result)
        except Exception as e:
            results.append({"symbol": pos["symbol"], "error": str(e), "signal": "CHECK_MANUALLY"})

    return results


def log_daily_prices():
    """Fikseeri hetkehinnad price history faili."""
    prices_data = check_prices()
    history = load_json(PRICES_FILE, {"entries": []})
    today = datetime.now().strftime("%Y-%m-%d")

    entry = {
        "date": today,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "positions": prices_data,
    }

    # Ära lisa duplikaati sama päeva kohta
    history["entries"] = [e for e in history["entries"] if e["date"] != today]
    history["entries"].append(entry)
    save_json(PRICES_FILE, history)
    return entry


def save_daily_snapshot():
    """Salvesta päeva portfelli snapshot."""
    prices = check_prices()
    snapshots = load_json(SNAPSHOTS_FILE, {"snapshots": []})
    today = datetime.now().strftime("%Y-%m-%d")

    total_invested = sum(p.get("amount_eur", 0) for p in prices if "amount_eur" in p)
    total_current = sum(p.get("current_value_eur", 0) for p in prices if "current_value_eur" in p)
    total_pnl = sum(p.get("unrealized_pnl_eur", 0) for p in prices if "unrealized_pnl_eur" in p)

    # Lisa ka realiseeritud kasum ja vaba raha
    journal = load_journal()
    config = load_config()
    realized_pnl = journal.get("stats", {}).get("total_profit_eur", 0)
    cash = calculate_available_cash()

    snapshot = {
        "date": today,
        "open_positions": len(prices),
        "available_cash_eur": cash,
        "total_invested_eur": round(total_invested, 2),
        "total_current_value_eur": round(total_current, 2),
        "total_portfolio_eur": round(cash + total_current, 2),
        "starting_capital_eur": config["starting_capital_eur"],
        "unrealized_pnl_eur": round(total_pnl, 2),
        "realized_pnl_eur": realized_pnl,
        "total_pnl_eur": round(total_pnl + realized_pnl, 2),
        "positions": [{
            "symbol": p["symbol"],
            "price": p.get("current_price", 0),
            "pct_change": p.get("pct_change", 0),
        } for p in prices if "symbol" in p],
    }

    # Ära lisa duplikaati sama päeva kohta
    snapshots["snapshots"] = [s for s in snapshots["snapshots"] if s["date"] != today]
    snapshots["snapshots"].append(snapshot)
    save_json(SNAPSHOTS_FILE, snapshots)
    return snapshot


# =============================================
# NÄDALA KOKKUVÕTE
# =============================================

def weekly_summary():
    """Genereeri nädala kokkuvõte."""
    journal = load_journal()
    prices = check_prices()
    config = load_config()

    today = datetime.now()
    week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    week_end = today.strftime("%Y-%m-%d")

    # Nädala tehingud
    week_trades = [
        t for t in journal["trades"]
        if t["date"][:10] >= week_start
    ]
    week_buys = [t for t in week_trades if t["action"] == "BUY"]
    week_sells = [t for t in week_trades if t["action"] == "SELL"]

    # Avatud positsioonid
    open_positions = []
    for p in prices:
        if "error" not in p:
            open_positions.append(p)

    # Nädala P&L
    week_realized = sum(t.get("profit_eur", 0) for t in week_sells if t.get("profit_eur"))
    total_unrealized = sum(p.get("unrealized_pnl_eur", 0) for p in open_positions)

    # Top/worst performer
    top_performer = max(open_positions, key=lambda p: p.get("pct_change", -999)) if open_positions else None
    worst_performer = min(open_positions, key=lambda p: p.get("pct_change", 999)) if open_positions else None

    cash = calculate_available_cash()

    summary = {
        "week": f"{week_start} - {week_end}",
        "week_start": week_start,
        "week_end": week_end,
        "generated": today.strftime("%Y-%m-%d %H:%M"),
        "available_cash_eur": cash,
        "starting_capital_eur": config["starting_capital_eur"],
        "trades_this_week": len(week_trades),
        "buys_this_week": len(week_buys),
        "sells_this_week": len(week_sells),
        "week_buys_detail": [{"symbol": t["symbol"], "price": t["price"], "amount": t["amount_eur"]} for t in week_buys],
        "week_sells_detail": [{"symbol": t["symbol"], "price": t["price"], "amount": t["amount_eur"], "pnl": t.get("profit_eur")} for t in week_sells],
        "week_realized_pnl_eur": round(week_realized, 2),
        "total_unrealized_pnl_eur": round(total_unrealized, 2),
        "open_positions": open_positions,
        "total_open": len(open_positions),
        "portfolio_value_eur": round(cash + sum(p.get("current_value_eur", 0) for p in open_positions), 2),
        "top_performer": {
            "symbol": top_performer["symbol"],
            "pct": top_performer["pct_change"],
        } if top_performer else None,
        "worst_performer": {
            "symbol": worst_performer["symbol"],
            "pct": worst_performer["pct_change"],
        } if worst_performer else None,
        "overall_stats": journal.get("stats", {}),
        "config": config,
    }

    # Salvesta weekly summaries faili
    weeklies = load_json(WEEKLY_FILE, {"summaries": []})
    weeklies["summaries"] = [s for s in weeklies["summaries"] if s.get("week_start") != week_start]
    weeklies["summaries"].append(summary)
    save_json(WEEKLY_FILE, weeklies)

    return summary


# =============================================
# EXCEL GENEREERIMINE
# =============================================

def generate_excel(output_path):
    """Genereeri täielik sandbox Excel raport."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference
    except ImportError:
        print("ERROR: openpyxl peab olema installeeritud.")
        print("Käivita: pip install openpyxl --break-system-packages")
        sys.exit(1)

    journal = load_journal()
    trades = journal.get("trades", [])
    stats = journal.get("stats", {})
    config = load_config()
    cash = calculate_available_cash()

    # OHUTUSKONTROLL: kui journal on tühi aga Excel fail on juba olemas tehingutega,
    # ära kirjuta üle (et mitte kaotada ajaloolisi andmeid kui JSON on kogemata kaduma läinud)
    if not trades and os.path.exists(output_path):
        import shutil
        backup_path = output_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        try:
            from openpyxl import load_workbook
            existing_wb = load_workbook(output_path, read_only=True)
            has_data = False
            if "Tehingud" in existing_wb.sheetnames:
                ws_check = existing_wb["Tehingud"]
                for row in ws_check.iter_rows(min_row=2, max_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        has_data = True
                        break
            existing_wb.close()
            if has_data:
                shutil.copy2(output_path, backup_path)
                print(f"⚠️  HOIATUS: Journal on tühi aga Excel sisaldab tehinguid!")
                print(f"   Backup salvestatud: {backup_path}")
                print(f"   KATKESTATUD: keeldun tühja andmetega üle kirjutamast.")
                print(f"   Kontrolli sandbox_journal.json — kas see puudub või on rikutud?")
                return
        except Exception as e:
            print(f"Hoiatus backup kontrollil: {e}")

    # Proovi laadida hindade ajalugu ja snapshotte
    price_history = load_json(PRICES_FILE, {"entries": []})
    snapshots = load_json(SNAPSHOTS_FILE, {"snapshots": []})
    weeklies = load_json(WEEKLY_FILE, {"summaries": []})

    # Stiilid
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BUY_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    SELL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    PROFIT_FONT = Font(name="Arial", color="2E7D32", bold=True)
    LOSS_FONT = Font(name="Arial", color="C62828", bold=True)
    TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E79")
    SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
    NORMAL_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    STAT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    SANDBOX_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    CASH_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    wb = Workbook()

    # ==========================================
    # SHEET 1: DASHBOARD
    # ==========================================
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells("A1:H1")
    ws["A1"] = "SANDBOX TRADER - Portfelli Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"SIMULEERITUD PORTFELL | Uuendatud: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, color="E65100", size=11)
    ws["A2"].fill = SANDBOX_FILL
    ws["A2"].alignment = Alignment(horizontal="center")

    # Kapitali ülevaade
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "KAPITALI ÜLEVAADE"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    open_trades = [t for t in trades if t.get("action") == "BUY" and t.get("status") == "OPEN"]
    invested = sum(t["amount_eur"] for t in open_trades)
    total_capital = cash + invested

    capital_data = [
        ("Stardikapital (€)", config["starting_capital_eur"]),
        ("Vaba raha (€)", cash),
        ("Investeeritud (€)", invested),
        ("Kogukapital (€)", total_capital),
        ("Kasum/Kahjum (€)", round(total_capital - config["starting_capital_eur"], 2)),
        ("Tootlus (%)", f"{round((total_capital / config['starting_capital_eur'] - 1) * 100, 2)}%"),
    ]

    for label, value in capital_data:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = BOLD_FONT
        ws[f"A{row}"].fill = CASH_FILL
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = NORMAL_FONT
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        if isinstance(value, (int, float)):
            if "kasum" in label.lower() or "kahjum" in label.lower():
                ws[f"B{row}"].font = PROFIT_FONT if value > 0 else LOSS_FONT if value < 0 else NORMAL_FONT
            if "vaba" in label.lower():
                ws[f"B{row}"].font = Font(name="Arial", bold=True, color="1565C0", size=10)
        row += 1

    # Statistika
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ÜLDSTATISTIKA"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    stat_data = [
        ("Kogu tehinguid", stats.get("total_trades", 0)),
        ("Avatud positsioonid", stats.get("open_positions", 0)),
        ("Suletud tehinguid", stats.get("closed_trades", 0)),
        ("Võitude arv", stats.get("wins", 0)),
        ("Kaotuste arv", stats.get("losses", 0)),
        ("Win Rate", f"{stats.get('win_rate', 0)}%"),
        ("Realiseeritud kasum (€)", stats.get("total_profit_eur", 0)),
        ("Keskmine kasum (%)", f"{stats.get('avg_profit_pct', 0)}%"),
    ]
    if stats.get("best_trade_pct"):
        stat_data.append(("Parim tehing", f"{stats.get('best_trade', '')} +{stats.get('best_trade_pct', 0)}%"))
        stat_data.append(("Halvim tehing", f"{stats.get('worst_trade', '')} {stats.get('worst_trade_pct', 0)}%"))

    for label, value in stat_data:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = BOLD_FONT
        ws[f"A{row}"].fill = STAT_FILL
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = NORMAL_FONT
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        if isinstance(value, (int, float)) and ("kasum" in label.lower() or "kahjum" in label.lower()):
            ws[f"B{row}"].font = PROFIT_FONT if value > 0 else LOSS_FONT if value < 0 else NORMAL_FONT
        row += 1

    # Avatud positsioonid
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "AVATUD POSITSIOONID"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    if open_trades:
        headers = ["Ticker", "Ostukuupäev", "Ostuhind ($)", "Aktsiad", "Summa (€)", "Analüüsi skoor", "Põhjus"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1
        for trade in open_trades:
            ws.cell(row=row, column=1, value=trade["symbol"]).font = BOLD_FONT
            ws.cell(row=row, column=2, value=trade["date"])
            ws.cell(row=row, column=3, value=trade["price"])
            ws.cell(row=row, column=4, value=trade["shares"])
            ws.cell(row=row, column=5, value=trade["amount_eur"])
            ws.cell(row=row, column=6, value=trade.get("analysis_score", "-"))
            ws.cell(row=row, column=7, value=trade.get("reason", ""))
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).fill = BUY_FILL
            row += 1
    else:
        ws[f"A{row}"] = "Hetkel avatud positsioone ei ole."
        ws[f"A{row}"].font = Font(name="Arial", italic=True, color="999999")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ==========================================
    # SHEET 2: TEHINGUD
    # ==========================================
    ws2 = wb.create_sheet("Tehingud")
    ws2.sheet_properties.tabColor = "2E7D32"

    headers = ["ID", "Kuupäev", "Ticker", "Tehing", "Aktsiad", "Hind ($)",
               "Summa (€)", "Skoor", "Põhjus", "Staatus", "Kasum (€)", "Kasum (%)"]

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i, trade in enumerate(trades, 2):
        ws2.cell(row=i, column=1, value=trade.get("id"))
        ws2.cell(row=i, column=2, value=trade.get("date"))
        ws2.cell(row=i, column=3, value=trade.get("symbol"))
        ws2.cell(row=i, column=4, value=trade.get("action"))
        ws2.cell(row=i, column=5, value=trade.get("shares"))
        ws2.cell(row=i, column=6, value=trade.get("price"))
        ws2.cell(row=i, column=7, value=trade.get("amount_eur"))
        ws2.cell(row=i, column=8, value=trade.get("analysis_score", "-"))
        ws2.cell(row=i, column=9, value=trade.get("reason", ""))
        ws2.cell(row=i, column=10, value=trade.get("status"))
        ws2.cell(row=i, column=11, value=trade.get("profit_eur"))
        ws2.cell(row=i, column=12, value=trade.get("profit_pct"))

        fill = BUY_FILL if trade.get("action") == "BUY" else SELL_FILL
        for col in range(1, 13):
            cell = ws2.cell(row=i, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.fill = fill

        if trade.get("profit_eur") is not None:
            pnl_font = PROFIT_FONT if trade["profit_eur"] > 0 else LOSS_FONT
            ws2.cell(row=i, column=11).font = pnl_font
            ws2.cell(row=i, column=12).font = pnl_font

    col_widths = [6, 18, 10, 10, 10, 12, 12, 8, 40, 10, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    if trades:
        ws2.auto_filter.ref = f"A1:L{len(trades) + 1}"

    # ==========================================
    # SHEET 3: PRICE TRACKING
    # ==========================================
    ws3 = wb.create_sheet("Price Tracking")
    ws3.sheet_properties.tabColor = "FF6F00"

    ws3["A1"] = "IGAPÄEVASED HINNAD"
    ws3["A1"].font = TITLE_FONT

    headers = ["Kuupäev", "Ticker", "Hind ($)", "Ostuhind ($)", "Muutus (%)", "Signaal"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 4
    for entry in price_history.get("entries", []):
        for pos in entry.get("positions", []):
            if "error" in pos:
                continue
            ws3.cell(row=row, column=1, value=entry["date"])
            ws3.cell(row=row, column=2, value=pos.get("symbol", ""))
            ws3.cell(row=row, column=3, value=pos.get("current_price", 0))
            ws3.cell(row=row, column=4, value=pos.get("buy_price", 0))
            ws3.cell(row=row, column=5, value=pos.get("pct_change", 0))
            ws3.cell(row=row, column=6, value=pos.get("signal", ""))
            for col in range(1, 7):
                ws3.cell(row=row, column=col).border = THIN_BORDER
                ws3.cell(row=row, column=col).font = NORMAL_FONT
            pct = pos.get("pct_change", 0)
            if pct > 0:
                ws3.cell(row=row, column=5).font = PROFIT_FONT
            elif pct < 0:
                ws3.cell(row=row, column=5).font = LOSS_FONT
            row += 1

    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    # ==========================================
    # SHEET 4: DAILY SNAPSHOTS
    # ==========================================
    ws4 = wb.create_sheet("Daily Snapshots")
    ws4.sheet_properties.tabColor = "4527A0"

    ws4["A1"] = "PORTFELLI PÄEVA SNAPSHOTS"
    ws4["A1"].font = TITLE_FONT

    headers = ["Kuupäev", "Avatud pos.", "Vaba raha (€)", "Investeeritud (€)", "Praegune väärtus (€)",
               "Koguportfell (€)", "Realiseerimata P&L (€)", "Realiseeritud P&L (€)", "Kogu P&L (€)"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 4
    for snap in snapshots.get("snapshots", []):
        ws4.cell(row=row, column=1, value=snap["date"])
        ws4.cell(row=row, column=2, value=snap.get("open_positions", 0))
        ws4.cell(row=row, column=3, value=snap.get("available_cash_eur", 0))
        ws4.cell(row=row, column=4, value=snap.get("total_invested_eur", 0))
        ws4.cell(row=row, column=5, value=snap.get("total_current_value_eur", 0))
        ws4.cell(row=row, column=6, value=snap.get("total_portfolio_eur", 0))
        ws4.cell(row=row, column=7, value=snap.get("unrealized_pnl_eur", 0))
        ws4.cell(row=row, column=8, value=snap.get("realized_pnl_eur", 0))
        ws4.cell(row=row, column=9, value=snap.get("total_pnl_eur", 0))
        for col in range(1, 10):
            ws4.cell(row=row, column=col).border = THIN_BORDER
            ws4.cell(row=row, column=col).font = NORMAL_FONT
        pnl = snap.get("total_pnl_eur", 0)
        if pnl > 0:
            ws4.cell(row=row, column=9).font = PROFIT_FONT
        elif pnl < 0:
            ws4.cell(row=row, column=9).font = LOSS_FONT
        row += 1

    # Lisa kumulatiivne P&L graafik kui piisavalt andmeid
    snap_list = snapshots.get("snapshots", [])
    if len(snap_list) >= 3:
        chart = LineChart()
        chart.title = "Portfelli väärtuse muutus"
        chart.y_axis.title = "Euro (€)"
        chart.style = 10
        chart.width = 22
        chart.height = 12
        data = Reference(ws4, min_col=6, min_row=3, max_row=3 + len(snap_list))
        cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(snap_list))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws4.add_chart(chart, "K3")

    for col in range(1, 10):
        ws4.column_dimensions[get_column_letter(col)].width = 20

    # ==========================================
    # SHEET 5: WEEKLY SUMMARY
    # ==========================================
    ws5 = wb.create_sheet("Weekly Summary")
    ws5.sheet_properties.tabColor = "00695C"

    ws5["A1"] = "NÄDALA KOKKUVÕTTED"
    ws5["A1"].font = TITLE_FONT

    headers = ["Nädal", "Tehinguid", "Ostud", "Müügid", "Nädala P&L (€)",
               "Realiseerimata P&L (€)", "Portfelli väärtus (€)", "Vaba raha (€)", "Top performer", "Worst performer"]
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 4
    for summary in weeklies.get("summaries", []):
        ws5.cell(row=row, column=1, value=summary.get("week", ""))
        ws5.cell(row=row, column=2, value=summary.get("trades_this_week", 0))
        ws5.cell(row=row, column=3, value=summary.get("buys_this_week", 0))
        ws5.cell(row=row, column=4, value=summary.get("sells_this_week", 0))
        ws5.cell(row=row, column=5, value=summary.get("week_realized_pnl_eur", 0))
        ws5.cell(row=row, column=6, value=summary.get("total_unrealized_pnl_eur", 0))
        ws5.cell(row=row, column=7, value=summary.get("portfolio_value_eur", 0))
        ws5.cell(row=row, column=8, value=summary.get("available_cash_eur", 0))
        top = summary.get("top_performer")
        worst = summary.get("worst_performer")
        ws5.cell(row=row, column=9, value=f"{top['symbol']} ({top['pct']:+.1f}%)" if top else "-")
        ws5.cell(row=row, column=10, value=f"{worst['symbol']} ({worst['pct']:+.1f}%)" if worst else "-")
        for col in range(1, 11):
            ws5.cell(row=row, column=col).border = THIN_BORDER
            ws5.cell(row=row, column=col).font = NORMAL_FONT
        row += 1

    for col in range(1, 11):
        ws5.column_dimensions[get_column_letter(col)].width = 20

    # ==========================================
    # SHEET 6: ANALÜÜS (Graafikud)
    # ==========================================
    ws6 = wb.create_sheet("Analüüs")
    ws6.sheet_properties.tabColor = "E65100"

    ws6["A1"] = "SANDBOX TEHINGUTE ANALÜÜS"
    ws6["A1"].font = TITLE_FONT

    closed_trades = [t for t in trades if t.get("profit_pct") is not None and t["action"] == "SELL"]

    if closed_trades:
        ws6["A3"] = "Tehing"
        ws6["B3"] = "Ticker"
        ws6["C3"] = "Kasum (%)"
        ws6["D3"] = "Kasum (€)"
        for col in range(1, 5):
            ws6.cell(row=3, column=col).font = HEADER_FONT
            ws6.cell(row=3, column=col).fill = HEADER_FILL

        for i, trade in enumerate(closed_trades, 4):
            ws6.cell(row=i, column=1, value=f"#{trade['id']}")
            ws6.cell(row=i, column=2, value=trade["symbol"])
            ws6.cell(row=i, column=3, value=trade["profit_pct"])
            ws6.cell(row=i, column=4, value=trade.get("profit_eur", 0))

        if len(closed_trades) >= 2:
            chart = BarChart()
            chart.title = "Kasum/Kahjum per tehing (%)"
            chart.y_axis.title = "Protsent (%)"
            chart.style = 10
            chart.width = 20
            chart.height = 12
            data = Reference(ws6, min_col=3, min_row=3, max_row=3 + len(closed_trades))
            cats = Reference(ws6, min_col=2, min_row=4, max_row=3 + len(closed_trades))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws6.add_chart(chart, "F3")
    else:
        ws6["A3"] = "Suletud tehinguid pole veel. Graafikud ilmuvad pärast esimest sandbox müüki."
        ws6["A3"].font = Font(name="Arial", italic=True, color="999999")

    for col in range(1, 7):
        ws6.column_dimensions[get_column_letter(col)].width = 16

    # ==========================================
    # SHEET 7: SEADED
    # ==========================================
    ws7 = wb.create_sheet("Seaded")
    ws7.sheet_properties.tabColor = "666666"

    ws7["A1"] = "SANDBOX TRADER SEADED"
    ws7["A1"].font = TITLE_FONT

    ws7["A2"] = "Režiim: SANDBOX (simuleeritud, tehinguid Lightyearis EI tehta)"
    ws7["A2"].font = Font(name="Arial", italic=True, color="E65100", size=11)
    ws7["A2"].fill = SANDBOX_FILL

    config_labels = {
        "starting_capital_eur": "Stardikapital (€)",
        "max_trade_eur": "Max tehing (€)",
        "target_profit_pct": "Kasumi sihtmärk (%)",
        "stop_loss_pct": "Stop-loss (%)",
        "max_open_positions": "Max avatud positsioonid",
        "risk_level": "Riskitase",
        "email": "Kokkuvõte email",
    }

    row = 4
    for key, label in config_labels.items():
        ws7[f"A{row}"] = label
        ws7[f"A{row}"].font = BOLD_FONT
        ws7[f"A{row}"].fill = STAT_FILL
        ws7[f"A{row}"].border = THIN_BORDER
        ws7[f"B{row}"] = config.get(key, "-")
        ws7[f"B{row}"].font = NORMAL_FONT
        ws7[f"B{row}"].border = THIN_BORDER
        row += 1

    # Lisa ka hetke rahajääk seadete lehele
    row += 1
    ws7[f"A{row}"] = "HETKE SEIS"
    ws7[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    cash_info = [
        ("Vaba raha (€)", cash),
        ("Investeeritud (€)", invested),
        ("Kogukapital (€)", total_capital),
    ]
    for label, value in cash_info:
        ws7[f"A{row}"] = label
        ws7[f"A{row}"].font = BOLD_FONT
        ws7[f"A{row}"].fill = CASH_FILL
        ws7[f"A{row}"].border = THIN_BORDER
        ws7[f"B{row}"] = value
        ws7[f"B{row}"].font = NORMAL_FONT
        ws7[f"B{row}"].border = THIN_BORDER
        row += 1

    ws7.column_dimensions["A"].width = 25
    ws7.column_dimensions["B"].width = 25

    wb.save(output_path)
    print(f"Sandbox Excel salvestatud: {output_path}")


# =============================================
# KONFIGURATSIOON
# =============================================

def set_config(key, value):
    config = load_config()
    try:
        value = int(value)
    except ValueError:
        try:
            value = float(value)
        except ValueError:
            pass
    config[key] = value
    save_json(CONFIG_FILE, config)
    print(f"Config updated: {key} = {value}")


# =============================================
# CLI
# =============================================

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "help"

    if cmd == "add_trade":
        # add_trade TICKER BUY/SELL SHARES PRICE AMOUNT_EUR "reason" [SCORE]
        if len(sys.argv) >= 8:
            trade = add_trade(
                sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5],
                sys.argv[6], sys.argv[7],
                int(sys.argv[8]) if len(sys.argv) > 8 else 0
            )
            print(json.dumps(trade, indent=2, default=str))
        else:
            print("Usage: add_trade TICKER BUY/SELL SHARES PRICE AMOUNT_EUR 'reason' [SCORE]")

    elif cmd == "check_prices":
        print(json.dumps(check_prices(), indent=2, default=str))

    elif cmd == "log_daily_prices":
        result = log_daily_prices()
        print(json.dumps(result, indent=2, default=str))

    elif cmd == "save_daily_snapshot":
        result = save_daily_snapshot()
        print(json.dumps(result, indent=2, default=str))

    elif cmd == "weekly_summary":
        result = weekly_summary()
        print(json.dumps(result, indent=2, default=str))

    elif cmd == "generate_excel":
        output = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SANDBOX_DIR, "sandbox_portfell.xlsx")
        generate_excel(output)

    elif cmd == "open_positions":
        print(json.dumps(get_open_positions(), indent=2, default=str))

    elif cmd == "stats":
        journal = load_journal()
        print(json.dumps(journal.get("stats", {}), indent=2))

    elif cmd == "config":
        print(json.dumps(load_config(), indent=2))

    elif cmd == "set_config":
        if len(sys.argv) >= 4:
            set_config(sys.argv[2], sys.argv[3])
        else:
            print("Usage: set_config KEY VALUE")

    elif cmd == "available_cash":
        cash = calculate_available_cash()
        config = load_config()
        print(json.dumps({
            "starting_capital_eur": config["starting_capital_eur"],
            "available_cash_eur": cash,
            "can_buy": cash > 0,
        }, indent=2))

    elif cmd == "summary":
        journal = load_journal()
        config = load_config()
        open_pos = get_open_positions()
        cash = calculate_available_cash()
        invested = round(sum(t["amount_eur"] for t in open_pos), 2)
        print(json.dumps({
            "mode": "SANDBOX",
            "config": config,
            "stats": journal.get("stats", {}),
            "capital": {
                "starting_capital_eur": config["starting_capital_eur"],
                "available_cash_eur": cash,
                "invested_eur": invested,
                "total_capital_eur": round(cash + invested, 2),
            },
            "open_positions": open_pos,
            "total_invested_eur": invested,
            "can_trade": len(open_pos) < config["max_open_positions"] and cash > 0,
        }, indent=2, default=str))

    elif cmd == "history":
        journal = load_journal()
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 20
        print(json.dumps(journal["trades"][-n:], indent=2, default=str))

    else:
        print("Sandbox Trader Commands:")
        print("  add_trade TICKER BUY/SELL SHARES PRICE AMOUNT_EUR 'reason' [SCORE]")
        print("  check_prices          - Kontrolli avatud positsioonide hetkehindu")
        print("  log_daily_prices      - Fikseeri hetkehinnad ajalukku")
        print("  save_daily_snapshot   - Salvesta päeva portfelli snapshot")
        print("  weekly_summary        - Genereeri nädala kokkuvõte")
        print("  generate_excel [PATH] - Genereeri Excel raport")
        print("  open_positions        - Näita avatud positsioone")
        print("  stats                 - Näita statistikat")
        print("  config                - Näita seadeid")
        print("  set_config KEY VALUE  - Muuda seadet")
        print("  available_cash        - Näita vaba raha")
        print("  summary               - Portfelli koondülevaade")
        print("  history [N]           - Viimased N tehingut")
