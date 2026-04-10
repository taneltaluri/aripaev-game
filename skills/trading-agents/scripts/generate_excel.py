#!/usr/bin/env python3
"""
Trading Journal Excel Generator - Genereerib ja uuendab Exceli faili tehingute jälgimiseks.
Loeb andmed trade_journal.json-ist ja loob professionaalse Excel raamatu.
"""
import json
import sys
import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("ERROR: openpyxl peab olema installeeritud.")
    print("Käivita: pip install openpyxl --break-system-packages")
    sys.exit(1)

JOURNAL_DIR = os.path.expanduser("~/.trading-agents")
JOURNAL_FILE = os.path.join(JOURNAL_DIR, "trade_journal.json")

# Värvid
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BUY_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SELL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
PROFIT_FONT = Font(name="Arial", color="2E7D32", bold=True)
LOSS_FONT = Font(name="Arial", color="C62828", bold=True)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
STAT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")


def load_journal() -> dict:
    if os.path.exists(JOURNAL_FILE):
        with open(JOURNAL_FILE, "r") as f:
            return json.load(f)
    return {"trades": [], "stats": {}}


def generate_excel(output_path: str):
    """Genereeri täielik Excel tehingute raport."""
    journal = load_journal()
    trades = journal.get("trades", [])
    stats = journal.get("stats", {})

    wb = Workbook()

    # ==========================================
    # SHEET 1: DASHBOARD (Kokkuvõte)
    # ==========================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "1F4E79"

    # Pealkiri
    ws_dash.merge_cells("A1:H1")
    ws_dash["A1"] = "Weekly Trader - Portfelli Dashboard"
    ws_dash["A1"].font = TITLE_FONT
    ws_dash["A1"].alignment = Alignment(horizontal="center")

    ws_dash.merge_cells("A2:H2")
    ws_dash["A2"] = f"Uuendatud: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_dash["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws_dash["A2"].alignment = Alignment(horizontal="center")

    # Statistika plokk
    stat_data = [
        ("Kogu tehinguid", stats.get("total_trades", 0)),
        ("Avatud positsioonid", stats.get("open_positions", 0)),
        ("Suletud tehinguid", stats.get("closed_trades", 0)),
        ("Võitude arv", stats.get("wins", 0)),
        ("Kaotuste arv", stats.get("losses", 0)),
        ("Win Rate", f"{stats.get('win_rate', 0)}%"),
        ("Kogukasum/-kahjum (€)", stats.get("total_profit_eur", 0)),
        ("Keskmine kasum (%)", f"{stats.get('avg_profit_pct', 0)}%"),
        ("Parim tehing (%)", f"+{stats.get('best_trade_pct', 0)}%"),
        ("Halvim tehing (%)", f"{stats.get('worst_trade_pct', 0)}%"),
    ]

    row = 4
    ws_dash.merge_cells(f"A{row}:D{row}")
    ws_dash[f"A{row}"] = "ÜLDSTATISTIKA"
    ws_dash[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    for label, value in stat_data:
        ws_dash[f"A{row}"] = label
        ws_dash[f"A{row}"].font = BOLD_FONT
        ws_dash[f"A{row}"].fill = STAT_FILL
        ws_dash[f"A{row}"].border = THIN_BORDER
        ws_dash[f"B{row}"] = value
        ws_dash[f"B{row}"].font = NORMAL_FONT
        ws_dash[f"B{row}"].border = THIN_BORDER
        ws_dash[f"B{row}"].alignment = Alignment(horizontal="right")

        if isinstance(value, (int, float)):
            if "kasum" in label.lower() or "kahjum" in label.lower():
                if isinstance(value, (int, float)) and value > 0:
                    ws_dash[f"B{row}"].font = PROFIT_FONT
                elif isinstance(value, (int, float)) and value < 0:
                    ws_dash[f"B{row}"].font = LOSS_FONT
        row += 1

    # Avatud positsioonid
    open_trades = [t for t in trades if t.get("action") == "BUY" and t.get("status") == "OPEN"]
    row += 2
    ws_dash.merge_cells(f"A{row}:F{row}")
    ws_dash[f"A{row}"] = "AVATUD POSITSIOONID"
    ws_dash[f"A{row}"].font = SUBTITLE_FONT
    row += 1

    if open_trades:
        headers = ["Ticker", "Ostukuupäev", "Ostuhind ($)", "Aktsiad", "Summa (€)", "Analüüsi skoor"]
        for col, header in enumerate(headers, 1):
            cell = ws_dash.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for trade in open_trades:
            ws_dash.cell(row=row, column=1, value=trade["symbol"]).font = BOLD_FONT
            ws_dash.cell(row=row, column=2, value=trade["date"])
            ws_dash.cell(row=row, column=3, value=trade["price"])
            ws_dash.cell(row=row, column=4, value=trade["shares"])
            ws_dash.cell(row=row, column=5, value=trade["amount_eur"])
            ws_dash.cell(row=row, column=6, value=trade.get("analysis_score", "-"))
            for col in range(1, 7):
                ws_dash.cell(row=row, column=col).border = THIN_BORDER
                ws_dash.cell(row=row, column=col).fill = BUY_FILL
            row += 1
    else:
        ws_dash[f"A{row}"] = "Hetkel avatud positsioone ei ole."
        ws_dash[f"A{row}"].font = Font(name="Arial", italic=True, color="999999")

    # Veergude laiused
    for col in range(1, 9):
        ws_dash.column_dimensions[get_column_letter(col)].width = 18

    # ==========================================
    # SHEET 2: KÕIK TEHINGUD (Detailne ajalugu)
    # ==========================================
    ws_trades = wb.create_sheet("Tehingud")
    ws_trades.sheet_properties.tabColor = "2E7D32"

    headers = [
        "ID", "Kuupäev", "Ticker", "Tehing", "Aktsiad", "Hind ($)",
        "Summa (€)", "Skoor", "Põhjus", "Staatus",
        "Kasum/Kahjum (€)", "Kasum/Kahjum (%)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_trades.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for i, trade in enumerate(trades, 2):
        ws_trades.cell(row=i, column=1, value=trade.get("id"))
        ws_trades.cell(row=i, column=2, value=trade.get("date"))
        ws_trades.cell(row=i, column=3, value=trade.get("symbol"))
        ws_trades.cell(row=i, column=4, value=trade.get("action"))
        ws_trades.cell(row=i, column=5, value=trade.get("shares"))
        ws_trades.cell(row=i, column=6, value=trade.get("price"))
        ws_trades.cell(row=i, column=7, value=trade.get("amount_eur"))
        ws_trades.cell(row=i, column=8, value=trade.get("analysis_score", "-"))
        ws_trades.cell(row=i, column=9, value=trade.get("reason", ""))
        ws_trades.cell(row=i, column=10, value=trade.get("status"))
        ws_trades.cell(row=i, column=11, value=trade.get("profit_eur"))
        ws_trades.cell(row=i, column=12, value=trade.get("profit_pct"))

        # Formattermine
        fill = BUY_FILL if trade.get("action") == "BUY" else SELL_FILL
        for col in range(1, 13):
            cell = ws_trades.cell(row=i, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.fill = fill

        # Kasum/kahjum värv
        if trade.get("profit_eur") is not None:
            pnl_cell = ws_trades.cell(row=i, column=11)
            pct_cell = ws_trades.cell(row=i, column=12)
            if trade["profit_eur"] > 0:
                pnl_cell.font = PROFIT_FONT
                pct_cell.font = PROFIT_FONT
            elif trade["profit_eur"] < 0:
                pnl_cell.font = LOSS_FONT
                pct_cell.font = LOSS_FONT

    # Veergude laiused
    col_widths = [6, 18, 10, 10, 10, 12, 12, 8, 40, 10, 16, 16]
    for i, width in enumerate(col_widths, 1):
        ws_trades.column_dimensions[get_column_letter(i)].width = width

    # Auto-filter
    ws_trades.auto_filter.ref = f"A1:L{len(trades) + 1}"

    # Kokku rida
    if trades:
        total_row = len(trades) + 2
        ws_trades.cell(row=total_row, column=6, value="KOKKU:").font = BOLD_FONT
        ws_trades.cell(row=total_row, column=7, value=f"=SUM(G2:G{len(trades)+1})")
        ws_trades.cell(row=total_row, column=7).font = BOLD_FONT
        ws_trades.cell(row=total_row, column=11, value=f"=SUM(K2:K{len(trades)+1})")
        ws_trades.cell(row=total_row, column=11).font = BOLD_FONT

    # ==========================================
    # SHEET 3: ANALÜÜS (Graafikud)
    # ==========================================
    ws_analysis = wb.create_sheet("Analüüs")
    ws_analysis.sheet_properties.tabColor = "E65100"

    ws_analysis["A1"] = "TEHINGUTE ANALÜÜS"
    ws_analysis["A1"].font = TITLE_FONT

    # Kasumi/kahjumi andmed graafikuks
    closed_trades = [t for t in trades if t.get("profit_pct") is not None]

    if closed_trades:
        ws_analysis["A3"] = "Tehing"
        ws_analysis["B3"] = "Ticker"
        ws_analysis["C3"] = "Kasum/Kahjum (%)"
        ws_analysis["D3"] = "Kasum/Kahjum (€)"
        for col in range(1, 5):
            ws_analysis.cell(row=3, column=col).font = HEADER_FONT
            ws_analysis.cell(row=3, column=col).fill = HEADER_FILL

        for i, trade in enumerate(closed_trades, 4):
            ws_analysis.cell(row=i, column=1, value=f"#{trade['id']}")
            ws_analysis.cell(row=i, column=2, value=trade["symbol"])
            ws_analysis.cell(row=i, column=3, value=trade["profit_pct"])
            ws_analysis.cell(row=i, column=4, value=trade.get("profit_eur", 0))

        # Kasumi graafik
        if len(closed_trades) >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Kasum/Kahjum per tehing (%)"
            chart.y_axis.title = "Protsent (%)"
            chart.x_axis.title = "Tehing"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data = Reference(ws_analysis, min_col=3, min_row=3, max_row=3 + len(closed_trades))
            cats = Reference(ws_analysis, min_col=2, min_row=4, max_row=3 + len(closed_trades))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws_analysis.add_chart(chart, "F3")

            # Kumulatiivne kasumi graafik
            cum_profit = 0
            ws_analysis[f"A{4 + len(closed_trades) + 2}"] = "Kumulatiivne kasum"
            ws_analysis[f"A{4 + len(closed_trades) + 2}"].font = SUBTITLE_FONT

            cum_row = 4 + len(closed_trades) + 3
            ws_analysis.cell(row=cum_row, column=1, value="Tehing #")
            ws_analysis.cell(row=cum_row, column=2, value="Kumulatiivne (€)")
            ws_analysis.cell(row=cum_row, column=1).font = HEADER_FONT
            ws_analysis.cell(row=cum_row, column=1).fill = HEADER_FILL
            ws_analysis.cell(row=cum_row, column=2).font = HEADER_FONT
            ws_analysis.cell(row=cum_row, column=2).fill = HEADER_FILL

            for i, trade in enumerate(closed_trades):
                cum_profit += trade.get("profit_eur", 0)
                ws_analysis.cell(row=cum_row + 1 + i, column=1, value=f"#{trade['id']}")
                ws_analysis.cell(row=cum_row + 1 + i, column=2, value=round(cum_profit, 2))

            if len(closed_trades) >= 2:
                line_chart = LineChart()
                line_chart.title = "Kumulatiivne kasum (€)"
                line_chart.y_axis.title = "Euro (€)"
                line_chart.style = 10
                line_chart.width = 20
                line_chart.height = 12

                data = Reference(ws_analysis, min_col=2, min_row=cum_row,
                               max_row=cum_row + len(closed_trades))
                cats = Reference(ws_analysis, min_col=1, min_row=cum_row + 1,
                               max_row=cum_row + len(closed_trades))
                line_chart.add_data(data, titles_from_data=True)
                line_chart.set_categories(cats)
                ws_analysis.add_chart(line_chart, f"F{cum_row}")

    else:
        ws_analysis["A3"] = "Suletud tehinguid pole veel. Graafikud ilmuvad pärast esimest müüki."
        ws_analysis["A3"].font = Font(name="Arial", italic=True, color="999999")

    # Veergude laiused
    for col in range(1, 7):
        ws_analysis.column_dimensions[get_column_letter(col)].width = 16

    # ==========================================
    # SHEET 4: SEADED
    # ==========================================
    ws_config = wb.create_sheet("Seaded")
    ws_config.sheet_properties.tabColor = "666666"

    ws_config["A1"] = "WEEKLY TRADER SEADED"
    ws_config["A1"].font = TITLE_FONT

    config_file = os.path.join(JOURNAL_DIR, "config.json")
    config = {}
    if os.path.exists(config_file):
        with open(config_file, "r") as f:
            config = json.load(f)

    config_labels = {
        "max_trade_eur": "Max tehing (€)",
        "target_profit_pct": "Kasumi sihtmärk (%)",
        "stop_loss_pct": "Stop-loss (%)",
        "max_open_positions": "Max avatud positsioonid",
        "weekly_trade_day": "Tehingupäev",
        "risk_level": "Riskitase",
    }

    row = 3
    for key, label in config_labels.items():
        ws_config[f"A{row}"] = label
        ws_config[f"A{row}"].font = BOLD_FONT
        ws_config[f"A{row}"].fill = STAT_FILL
        ws_config[f"A{row}"].border = THIN_BORDER
        ws_config[f"B{row}"] = config.get(key, "-")
        ws_config[f"B{row}"].font = NORMAL_FONT
        ws_config[f"B{row}"].border = THIN_BORDER
        row += 1

    ws_config.column_dimensions["A"].width = 25
    ws_config.column_dimensions["B"].width = 15

    # Salvesta
    wb.save(output_path)
    print(f"Excel salvestatud: {output_path}")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else os.path.join(JOURNAL_DIR, "weekly_trader_journal.xlsx")
    generate_excel(output)
